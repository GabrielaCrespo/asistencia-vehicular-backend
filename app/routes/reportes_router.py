from fastapi import APIRouter, HTTPException, Header, Query, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from psycopg2.extras import RealDictCursor
from typing import Optional, List, Any, Dict
import csv
import io
import json
from datetime import datetime

from ..classes.postgresql import Database
from ..utils.tenant_deps import get_token_payload, assert_taller_access, assert_org_access

router = APIRouter(prefix="/api/reportes", tags=["Reportes"])


# ===================== HELPERS MULTI-TENANT =====================

def _taller_ids_para_payload(payload: dict, db, taller_id: int = None, org_id: int = None) -> tuple[str, list]:
    """
    Returns a SQL fragment (with %s placeholders) that can be appended after WHERE
    to enforce multi-tenant isolation, plus the list of params.

    Returned format:  ("a.taller_id IN %s", [(1,2,3)])
    Or for single:    ("a.taller_id = %s", [7])
    """
    rol = payload.get("rol")

    if rol == "administrador":
        if taller_id:
            return "a.taller_id = %s", [taller_id]
        if org_id:
            return "t.organizacion_id = %s", [org_id]
        return "1=1", []

    if rol == "tenant_admin":
        effective_org = org_id or payload.get("organizacion_id")
        if org_id:
            assert_org_access(payload, org_id)
        return "t.organizacion_id = %s", [effective_org]

    if rol == "taller":
        effective_taller = taller_id or payload.get("taller_id")
        if taller_id:
            assert_taller_access(payload, taller_id, db)
        return "a.taller_id = %s", [effective_taller]

    raise HTTPException(status_code=403, detail="Rol no autorizado para reportes")


def _calificacion_tenant_filter(payload: dict, db, taller_id: int = None, org_id: int = None) -> tuple[str, list]:
    """Same as above but for CALIFICACION table (no asignacion join needed)."""
    rol = payload.get("rol")

    if rol == "administrador":
        if taller_id:
            return "c.taller_id = %s", [taller_id]
        if org_id:
            return "t.organizacion_id = %s", [org_id]
        return "1=1", []

    if rol == "tenant_admin":
        effective_org = org_id or payload.get("organizacion_id")
        if org_id:
            assert_org_access(payload, org_id)
        return "t.organizacion_id = %s", [effective_org]

    if rol == "taller":
        effective_taller = taller_id or payload.get("taller_id")
        if taller_id:
            assert_taller_access(payload, taller_id, db)
        return "c.taller_id = %s", [effective_taller]

    raise HTTPException(status_code=403, detail="Rol no autorizado para reportes")


def _add_date_filter(conditions: list, params: list, fecha_desde: str = None, fecha_hasta: str = None, col: str = "i.fecha_creacion"):
    if fecha_desde:
        conditions.append(f"{col} >= %s")
        params.append(fecha_desde)
    if fecha_hasta:
        conditions.append(f"{col} <= %s")
        params.append(fecha_hasta + " 23:59:59")


# ===================== MODELOS REQUEST =====================

class ReporteDinamicoFiltros(BaseModel):
    fecha_desde: Optional[str] = None
    fecha_hasta: Optional[str] = None
    taller_id: Optional[int] = None
    org_id: Optional[int] = None
    tipo_incidente: Optional[str] = None
    estado: Optional[str] = None
    tecnico_id: Optional[int] = None
    zona: Optional[str] = None


class VozRequest(BaseModel):
    texto: str
    taller_id: Optional[int] = None
    org_id: Optional[int] = None


# ===================== QUERIES COMPARTIDAS =====================

_EMERGENCIAS_SELECT = """
    SELECT
        i.incidente_id,
        i.tipo_problema,
        i.descripcion,
        i.estado AS estado_incidente,
        i.prioridad,
        i.fecha_creacion,
        i.fecha_cierre,
        a.asignacion_id,
        a.estado AS estado_asignacion,
        a.fecha_asignacion,
        a.fecha_inicio_servicio,
        a.fecha_cierre_servicio,
        COALESCE(EXTRACT(EPOCH FROM (a.fecha_cierre_servicio - a.fecha_asignacion))/60, 0)::int AS duracion_min,
        t.razon_social AS taller,
        u.nombre AS cliente,
        u.telefono AS cliente_telefono,
        COALESCE(v.marca || ' ' || v.modelo, 'Sin vehículo') AS vehiculo,
        v.placa,
        COALESCE(te.nombre, 'Sin asignar') AS tecnico,
        COALESCE(p.monto_total, 0) AS monto_total
    FROM incidente i
    JOIN asignacion a ON a.incidente_id = i.incidente_id
    JOIN taller t ON t.taller_id = a.taller_id
    JOIN usuario u ON u.usuario_id = i.usuario_id
    LEFT JOIN vehiculo v ON v.vehiculo_id = i.vehiculo_id
    LEFT JOIN tecnico te ON te.tecnico_id = a.tecnico_id
    LEFT JOIN pago p ON p.incidente_id = i.incidente_id
"""

_HISTORIAL_SELECT = """
    SELECT
        a.asignacion_id,
        i.incidente_id,
        i.tipo_problema,
        i.descripcion,
        a.estado,
        i.prioridad,
        i.fecha_creacion,
        a.fecha_asignacion,
        a.fecha_inicio_servicio,
        a.fecha_cierre_servicio,
        COALESCE(EXTRACT(EPOCH FROM (a.fecha_cierre_servicio - a.fecha_inicio_servicio))/60, 0)::int AS duracion_servicio_min,
        t.razon_social AS taller,
        u.nombre AS cliente,
        COALESCE(v.marca || ' ' || v.modelo, 'Sin vehículo') AS vehiculo,
        v.placa,
        COALESCE(te.nombre, 'Sin asignar') AS tecnico,
        COALESCE(p.monto_total, 0) AS monto_total,
        COALESCE(p.monto_taller, 0) AS monto_taller,
        COALESCE(p.comision_plataforma, 0) AS comision,
        COALESCE(cal.puntuacion, 0) AS calificacion
    FROM asignacion a
    JOIN incidente i ON i.incidente_id = a.incidente_id
    JOIN taller t ON t.taller_id = a.taller_id
    JOIN usuario u ON u.usuario_id = i.usuario_id
    LEFT JOIN vehiculo v ON v.vehiculo_id = i.vehiculo_id
    LEFT JOIN tecnico te ON te.tecnico_id = a.tecnico_id
    LEFT JOIN pago p ON p.incidente_id = i.incidente_id
    LEFT JOIN calificacion cal ON cal.incidente_id = i.incidente_id
"""

_INGRESOS_SELECT = """
    SELECT
        p.pago_id,
        p.fecha_pago,
        p.monto_total,
        p.monto_servicio,
        p.monto_taller,
        p.comision_plataforma,
        p.metodo_pago,
        p.estado,
        t.razon_social AS taller,
        i.tipo_problema,
        u.nombre AS cliente
    FROM pago p
    JOIN incidente i ON i.incidente_id = p.incidente_id
    JOIN asignacion a ON a.incidente_id = i.incidente_id AND a.taller_id = (
        SELECT taller_id FROM asignacion WHERE incidente_id = i.incidente_id
        ORDER BY fecha_asignacion DESC LIMIT 1
    )
    JOIN taller t ON t.taller_id = a.taller_id
    JOIN usuario u ON u.usuario_id = i.usuario_id
"""

_SLA_SELECT = """
    SELECT
        t.taller_id,
        t.razon_social AS taller,
        COUNT(*) AS total_servicios,
        SUM(CASE WHEN a.estado = 'completada' THEN 1 ELSE 0 END) AS completados,
        ROUND(AVG(
            EXTRACT(EPOCH FROM (a.fecha_inicio_servicio - a.fecha_asignacion))/60
        )::numeric, 1) AS tiempo_respuesta_prom_min,
        ROUND(AVG(
            EXTRACT(EPOCH FROM (a.fecha_cierre_servicio - a.fecha_inicio_servicio))/60
        )::numeric, 1) AS duracion_servicio_prom_min,
        SUM(CASE WHEN EXTRACT(EPOCH FROM (a.fecha_inicio_servicio - a.fecha_asignacion))/60 <= 30
            THEN 1 ELSE 0 END) AS dentro_sla,
        ROUND(
            100.0 * SUM(CASE WHEN EXTRACT(EPOCH FROM (a.fecha_inicio_servicio - a.fecha_asignacion))/60 <= 30
                THEN 1 ELSE 0 END) / NULLIF(COUNT(*), 0), 1
        ) AS cumplimiento_sla_pct
    FROM asignacion a
    JOIN taller t ON t.taller_id = a.taller_id
"""


COLUMNAS_EXPORT: dict = {
    'emergencias':         ['fecha_creacion', 'tipo_problema', 'estado_asignacion', 'prioridad', 'cliente', 'vehiculo', 'taller', 'tecnico', 'monto_total'],
    'historial-servicios': ['fecha_asignacion', 'tipo_problema', 'estado', 'cliente', 'vehiculo', 'taller', 'tecnico', 'monto_total', 'calificacion'],
    'ingresos':            ['fecha_pago', 'taller', 'cliente', 'tipo_problema', 'monto_total', 'monto_taller', 'comision_plataforma', 'metodo_pago'],
    'calificaciones':      ['fecha_calificacion', 'taller', 'cliente', 'tipo_problema', 'puntuacion', 'puntuacion_servicio', 'comentario'],
    'kpis':                ['taller', 'total_servicios', 'completados', 'pendientes', 'urgentes', 'ingresos_netos', 'calificacion_promedio', 'tiempo_respuesta_prom'],
    'incidentes-tipo':     ['tipo_problema', 'total', 'completados', 'urgentes', 'ticket_promedio'],
    'sla':                 ['taller', 'total_servicios', 'completados', 'tiempo_respuesta_prom_min', 'cumplimiento_sla_pct'],
    'dinamico':            ['fecha_asignacion', 'tipo_problema', 'estado', 'cliente', 'vehiculo', 'taller', 'tecnico', 'monto_total', 'calificacion'],
}

COL_LABELS_EXPORT: dict = {
    'fecha_creacion': 'Fecha', 'fecha_asignacion': 'Fecha', 'fecha_pago': 'Fecha Pago',
    'fecha_calificacion': 'Fecha', 'tipo_problema': 'Tipo Incidente',
    'estado_asignacion': 'Estado', 'estado': 'Estado', 'prioridad': 'Prioridad',
    'cliente': 'Cliente', 'vehiculo': 'Vehículo', 'taller': 'Taller', 'tecnico': 'Técnico',
    'monto_total': 'Monto Total', 'monto_taller': 'Monto Taller',
    'comision_plataforma': 'Comisión', 'metodo_pago': 'Método Pago',
    'puntuacion': 'Puntuación', 'puntuacion_servicio': 'Punt. Servicio',
    'comentario': 'Comentario', 'total_servicios': 'Total', 'completados': 'Completados',
    'pendientes': 'Pendientes', 'urgentes': 'Urgentes', 'ingresos_netos': 'Ingresos Netos',
    'calificacion_promedio': 'Cal. Prom.', 'calificacion': 'Calificación',
    'tiempo_respuesta_prom': 'T. Resp. (min)', 'tiempo_respuesta_prom_min': 'T. Resp. (min)',
    'cumplimiento_sla_pct': 'SLA %', 'ticket_promedio': 'Ticket Prom.', 'total': 'Total',
}


def _filter_rows(rows: list, tipo: str) -> tuple:
    """Returns (filtered_rows, col_keys, col_labels) with only the curated columns."""
    cols = COLUMNAS_EXPORT.get(tipo, list(rows[0].keys()) if rows else [])
    if rows:
        available = set(rows[0].keys())
        cols = [c for c in cols if c in available]
        filtered = [{c: r.get(c) for c in cols} for r in rows]
    else:
        filtered = []
    labels = [COL_LABELS_EXPORT.get(c, c.replace('_', ' ').title()) for c in cols]
    return filtered, cols, labels


def _rows_to_dicts(rows) -> list:
    result = []
    for row in rows:
        d = {}
        for k, v in row.items():
            if hasattr(v, "isoformat"):
                d[k] = v.isoformat()
            elif v is None:
                d[k] = None
            else:
                d[k] = v
        result.append(d)
    return result


# ===================== ENDPOINTS ESTÁTICOS =====================

@router.get("/emergencias")
def reporte_emergencias(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    taller_id: Optional[int] = Query(None),
    org_id: Optional[int] = Query(None),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    payload = get_token_payload(authorization)
    tenant_cond, tenant_params = _taller_ids_para_payload(payload, db, taller_id, org_id)

    conditions = [tenant_cond]
    params = tenant_params[:]
    _add_date_filter(conditions, params, fecha_desde, fecha_hasta)

    where = " WHERE " + " AND ".join(conditions)
    query = _EMERGENCIAS_SELECT + where + " ORDER BY i.fecha_creacion DESC"

    cur = db.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(query, params)
        rows = cur.fetchall()
    finally:
        cur.close()

    data = _rows_to_dicts(rows)
    total = len(data)
    completadas = sum(1 for r in data if r.get("estado_asignacion") == "completada")
    monto_total = sum(r.get("monto_total") or 0 for r in data)

    return {
        "resumen": {"total": total, "completadas": completadas, "monto_total": round(monto_total, 2)},
        "datos": data,
    }


@router.get("/historial-servicios")
def reporte_historial(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    taller_id: Optional[int] = Query(None),
    org_id: Optional[int] = Query(None),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    payload = get_token_payload(authorization)
    tenant_cond, tenant_params = _taller_ids_para_payload(payload, db, taller_id, org_id)

    conditions = [tenant_cond]
    params = tenant_params[:]
    _add_date_filter(conditions, params, fecha_desde, fecha_hasta)

    where = " WHERE " + " AND ".join(conditions)
    query = _HISTORIAL_SELECT + where + " ORDER BY a.fecha_asignacion DESC"

    cur = db.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(query, params)
        rows = cur.fetchall()
    finally:
        cur.close()

    data = _rows_to_dicts(rows)
    total = len(data)
    monto = sum(r.get("monto_total") or 0 for r in data)
    cal_list = [r["calificacion"] for r in data if r.get("calificacion")]
    promedio_cal = round(sum(cal_list) / len(cal_list), 2) if cal_list else 0

    return {
        "resumen": {"total_servicios": total, "ingresos_totales": round(monto, 2), "calificacion_promedio": promedio_cal},
        "datos": data,
    }


@router.get("/ingresos")
def reporte_ingresos(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    taller_id: Optional[int] = Query(None),
    org_id: Optional[int] = Query(None),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    payload = get_token_payload(authorization)
    tenant_cond, tenant_params = _taller_ids_para_payload(payload, db, taller_id, org_id)

    conditions = [tenant_cond, "p.estado = 'completado'"]
    params = tenant_params[:]
    _add_date_filter(conditions, params, fecha_desde, fecha_hasta, col="p.fecha_pago")

    where = " WHERE " + " AND ".join(conditions)
    query = _INGRESOS_SELECT + where + " ORDER BY p.fecha_pago DESC"

    cur = db.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(query, params)
        rows = cur.fetchall()
    finally:
        cur.close()

    data = _rows_to_dicts(rows)
    bruto = sum(r.get("monto_total") or 0 for r in data)
    neto = sum(r.get("monto_taller") or 0 for r in data)
    comisiones = sum(r.get("comision_plataforma") or 0 for r in data)

    return {
        "resumen": {
            "total_transacciones": len(data),
            "ingresos_brutos": round(bruto, 2),
            "ingresos_netos": round(neto, 2),
            "comisiones": round(comisiones, 2),
        },
        "datos": data,
    }


@router.get("/calificaciones")
def reporte_calificaciones(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    taller_id: Optional[int] = Query(None),
    org_id: Optional[int] = Query(None),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    payload = get_token_payload(authorization)
    tenant_cond, tenant_params = _calificacion_tenant_filter(payload, db, taller_id, org_id)

    conditions = [tenant_cond]
    params = tenant_params[:]
    _add_date_filter(conditions, params, fecha_desde, fecha_hasta, col="c.fecha_calificacion")

    where = " WHERE " + " AND ".join(conditions)
    query = f"""
        SELECT
            c.calificacion_id,
            c.puntuacion,
            c.puntuacion_servicio,
            c.aspecto_atencion,
            c.aspecto_puntualidad,
            c.aspecto_limpieza,
            c.comentario,
            c.fecha_calificacion,
            t.razon_social AS taller,
            u.nombre AS cliente,
            i.tipo_problema
        FROM calificacion c
        JOIN taller t ON t.taller_id = c.taller_id
        JOIN usuario u ON u.usuario_id = c.usuario_id
        JOIN incidente i ON i.incidente_id = c.incidente_id
        {where}
        ORDER BY c.fecha_calificacion DESC
    """

    cur = db.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(query, params)
        rows = cur.fetchall()
    finally:
        cur.close()

    data = _rows_to_dicts(rows)
    puntuaciones = [r["puntuacion"] for r in data if r.get("puntuacion")]
    promedio = round(sum(puntuaciones) / len(puntuaciones), 2) if puntuaciones else 0

    return {
        "resumen": {
            "total_calificaciones": len(data),
            "promedio_general": promedio,
            "distribucion": {str(i): sum(1 for p in puntuaciones if p == i) for i in range(1, 6)},
        },
        "datos": data,
    }


@router.get("/kpis")
def reporte_kpis(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    taller_id: Optional[int] = Query(None),
    org_id: Optional[int] = Query(None),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    payload = get_token_payload(authorization)
    tenant_cond, tenant_params = _taller_ids_para_payload(payload, db, taller_id, org_id)

    conditions = [tenant_cond]
    params = tenant_params[:]
    _add_date_filter(conditions, params, fecha_desde, fecha_hasta)

    where = " WHERE " + " AND ".join(conditions)

    query = f"""
        SELECT
            t.taller_id,
            t.razon_social AS taller,
            COUNT(a.asignacion_id) AS total_servicios,
            SUM(CASE WHEN a.estado = 'completada' THEN 1 ELSE 0 END) AS completados,
            SUM(CASE WHEN a.estado = 'pendiente' THEN 1 ELSE 0 END) AS pendientes,
            SUM(CASE WHEN i.prioridad = 'urgente' THEN 1 ELSE 0 END) AS urgentes,
            COALESCE(AVG(EXTRACT(EPOCH FROM (a.fecha_inicio_servicio - a.fecha_asignacion))/60), 0)::int AS tiempo_respuesta_prom,
            COALESCE(SUM(p.monto_taller), 0) AS ingresos_netos,
            COALESCE(AVG(cal.puntuacion), 0)::numeric(3,1) AS calificacion_promedio,
            COALESCE(SUM(p.comision_plataforma), 0) AS comisiones_generadas
        FROM asignacion a
        JOIN taller t ON t.taller_id = a.taller_id
        JOIN incidente i ON i.incidente_id = a.incidente_id
        LEFT JOIN pago p ON p.incidente_id = i.incidente_id
        LEFT JOIN calificacion cal ON cal.incidente_id = i.incidente_id
        {where}
        GROUP BY t.taller_id, t.razon_social
        ORDER BY completados DESC
    """

    cur = db.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(query, params)
        rows = cur.fetchall()
    finally:
        cur.close()

    data = _rows_to_dicts(rows)
    total_s = sum(r.get("total_servicios") or 0 for r in data)
    total_c = sum(r.get("completados") or 0 for r in data)
    tasa = round(total_c / total_s * 100, 1) if total_s else 0

    return {
        "resumen": {
            "total_servicios": total_s,
            "completados": total_c,
            "tasa_completacion_pct": tasa,
            "ingresos_totales": round(sum(float(r.get("ingresos_netos") or 0) for r in data), 2),
        },
        "por_taller": data,
    }


@router.get("/incidentes-tipo")
def reporte_incidentes_tipo(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    taller_id: Optional[int] = Query(None),
    org_id: Optional[int] = Query(None),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    payload = get_token_payload(authorization)
    tenant_cond, tenant_params = _taller_ids_para_payload(payload, db, taller_id, org_id)

    conditions = [tenant_cond]
    params = tenant_params[:]
    _add_date_filter(conditions, params, fecha_desde, fecha_hasta)

    where = " WHERE " + " AND ".join(conditions)

    query = f"""
        SELECT
            COALESCE(i.tipo_problema, 'Sin clasificar') AS tipo_problema,
            COUNT(*) AS total,
            SUM(CASE WHEN a.estado = 'completada' THEN 1 ELSE 0 END) AS completados,
            SUM(CASE WHEN i.prioridad = 'urgente' THEN 1 ELSE 0 END) AS urgentes,
            ROUND(AVG(p.monto_total)::numeric, 2) AS ticket_promedio
        FROM incidente i
        JOIN asignacion a ON a.incidente_id = i.incidente_id
        JOIN taller t ON t.taller_id = a.taller_id
        LEFT JOIN pago p ON p.incidente_id = i.incidente_id
        {where}
        GROUP BY i.tipo_problema
        ORDER BY total DESC
    """

    cur = db.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(query, params)
        rows = cur.fetchall()
    finally:
        cur.close()

    data = _rows_to_dicts(rows)
    return {"datos": data}


@router.get("/sla")
def reporte_sla(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    taller_id: Optional[int] = Query(None),
    org_id: Optional[int] = Query(None),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    payload = get_token_payload(authorization)
    tenant_cond, tenant_params = _taller_ids_para_payload(payload, db, taller_id, org_id)

    conditions = [tenant_cond]
    params = tenant_params[:]
    _add_date_filter(conditions, params, fecha_desde, fecha_hasta)

    where = " WHERE " + " AND ".join(conditions)
    query = _SLA_SELECT + where + " GROUP BY t.taller_id, t.razon_social ORDER BY cumplimiento_sla_pct DESC NULLS LAST"

    cur = db.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(query, params)
        rows = cur.fetchall()
    finally:
        cur.close()

    data = _rows_to_dicts(rows)
    cumplimientos = [float(r["cumplimiento_sla_pct"] or 0) for r in data]
    prom_sla = round(sum(cumplimientos) / len(cumplimientos), 1) if cumplimientos else 0

    return {
        "resumen": {"promedio_cumplimiento_sla_pct": prom_sla},
        "por_taller": data,
    }


# ===================== REPORTE DINÁMICO =====================

@router.post("/dinamico")
def reporte_dinamico(
    filtros: ReporteDinamicoFiltros,
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    payload = get_token_payload(authorization)
    tenant_cond, tenant_params = _taller_ids_para_payload(
        payload, db, filtros.taller_id, filtros.org_id
    )

    conditions = [tenant_cond]
    params = tenant_params[:]
    _add_date_filter(conditions, params, filtros.fecha_desde, filtros.fecha_hasta)

    if filtros.tipo_incidente:
        conditions.append("i.tipo_problema ILIKE %s")
        params.append(f"%{filtros.tipo_incidente}%")

    if filtros.estado:
        conditions.append("a.estado = %s")
        params.append(filtros.estado)

    if filtros.tecnico_id:
        conditions.append("a.tecnico_id = %s")
        params.append(filtros.tecnico_id)

    if filtros.zona:
        conditions.append("(t.direccion ILIKE %s OR t.razon_social ILIKE %s)")
        params.extend([f"%{filtros.zona}%", f"%{filtros.zona}%"])

    where = " WHERE " + " AND ".join(conditions)
    query = _HISTORIAL_SELECT + where + " ORDER BY a.fecha_asignacion DESC LIMIT 1000"

    cur = db.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(query, params)
        rows = cur.fetchall()
    finally:
        cur.close()

    data = _rows_to_dicts(rows)
    monto = sum(r.get("monto_total") or 0 for r in data)
    cal_list = [r["calificacion"] for r in data if r.get("calificacion")]

    return {
        "filtros_aplicados": filtros.model_dump(exclude_none=True),
        "resumen": {
            "total_registros": len(data),
            "ingresos_totales": round(monto, 2),
            "calificacion_promedio": round(sum(cal_list) / len(cal_list), 2) if cal_list else 0,
        },
        "datos": data,
    }


# ===================== REPORTE POR VOZ =====================

_INTENT_PROMPT = """Eres un asistente de reportes para una plataforma de asistencia vehicular.
El usuario ha dicho (en texto): "{texto}"

Identifica el tipo de reporte y extrae filtros. Responde ÚNICAMENTE en JSON con este esquema:
{{
  "tipo_reporte": "emergencias|historial-servicios|ingresos|calificaciones|kpis|incidentes-tipo|sla|dinamico",
  "filtros": {{
    "fecha_desde": "YYYY-MM-DD o null",
    "fecha_hasta": "YYYY-MM-DD o null",
    "tipo_incidente": "texto o null",
    "estado": "pendiente|asignado|en_camino|en_servicio|completada o null",
    "formato_exportar": "csv|pdf|excel o null"
  }},
  "mensaje_confirmacion": "frase corta confirmando lo que vas a generar"
}}

Reglas para inferir fechas relativas (hoy es {hoy}):
- "último mes" → fecha_desde = primer día del mes pasado, fecha_hasta = hoy
- "esta semana" → fecha_desde = lunes de esta semana
- "hoy" → fecha_desde = hoy
- "último trimestre" → fecha_desde = hace 90 días

Si no puedes determinar el tipo de reporte, usa "dinamico".
No incluyas ningún texto fuera del JSON."""


@router.post("/voz")
def reporte_voz(
    body: VozRequest,
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    from ..services.config import Config
    import google.generativeai as genai

    payload = get_token_payload(authorization)

    try:
        genai.configure(api_key=Config.GEMINI_API_KEY)
        model = genai.GenerativeModel("gemini-2.5-flash")
        hoy = datetime.now().strftime("%Y-%m-%d")
        prompt = _INTENT_PROMPT.format(texto=body.texto, hoy=hoy)
        response = model.generate_content(prompt)
        raw = response.text.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        interpretacion = json.loads(raw)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo interpretar la consulta: {str(e)}")

    tipo = interpretacion.get("tipo_reporte", "dinamico")
    filtros_raw = interpretacion.get("filtros", {})
    mensaje = interpretacion.get("mensaje_confirmacion", "Generando reporte...")

    fecha_desde = filtros_raw.get("fecha_desde")
    fecha_hasta = filtros_raw.get("fecha_hasta")
    tipo_incidente = filtros_raw.get("tipo_incidente")
    estado = filtros_raw.get("estado")

    dispatch = {
        "emergencias": lambda: reporte_emergencias(
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            taller_id=body.taller_id, org_id=body.org_id,
            authorization=authorization, db=db
        ),
        "historial-servicios": lambda: reporte_historial(
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            taller_id=body.taller_id, org_id=body.org_id,
            authorization=authorization, db=db
        ),
        "ingresos": lambda: reporte_ingresos(
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            taller_id=body.taller_id, org_id=body.org_id,
            authorization=authorization, db=db
        ),
        "calificaciones": lambda: reporte_calificaciones(
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            taller_id=body.taller_id, org_id=body.org_id,
            authorization=authorization, db=db
        ),
        "kpis": lambda: reporte_kpis(
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            taller_id=body.taller_id, org_id=body.org_id,
            authorization=authorization, db=db
        ),
        "incidentes-tipo": lambda: reporte_incidentes_tipo(
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            taller_id=body.taller_id, org_id=body.org_id,
            authorization=authorization, db=db
        ),
        "sla": lambda: reporte_sla(
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            taller_id=body.taller_id, org_id=body.org_id,
            authorization=authorization, db=db
        ),
    }

    if tipo in dispatch:
        resultado = dispatch[tipo]()
    else:
        filtros_din = ReporteDinamicoFiltros(
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            taller_id=body.taller_id,
            org_id=body.org_id,
            tipo_incidente=tipo_incidente,
            estado=estado,
        )
        resultado = reporte_dinamico(filtros_din, authorization=authorization, db=db)

    return {
        "tipo_reporte": tipo,
        "mensaje_confirmacion": mensaje,
        "formato_exportar": filtros_raw.get("formato_exportar"),
        "filtros_aplicados": {
            "fecha_desde": fecha_desde,
            "fecha_hasta": fecha_hasta,
            "tipo_incidente": tipo_incidente,
            "estado": estado,
        },
        "resultado": resultado,
    }


# ===================== EXPORTACIÓN =====================

def _build_export_data(tipo: str, fecha_desde, fecha_hasta, taller_id, org_id, authorization, db) -> tuple[list, str]:
    """Returns (rows_list, titulo_reporte)."""
    dispatch = {
        "emergencias": (reporte_emergencias, "Emergencias Atendidas"),
        "historial-servicios": (reporte_historial, "Historial de Servicios"),
        "ingresos": (reporte_ingresos, "Ingresos Generados"),
        "calificaciones": (reporte_calificaciones, "Calificaciones de Talleres"),
        "kpis": (reporte_kpis, "KPIs Operacionales"),
        "incidentes-tipo": (reporte_incidentes_tipo, "Incidentes por Tipo"),
        "sla": (reporte_sla, "Cumplimiento SLA"),
    }

    if tipo not in dispatch:
        raise HTTPException(status_code=400, detail=f"Tipo de reporte desconocido: {tipo}")

    fn, titulo = dispatch[tipo]
    result = fn(
        fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
        taller_id=taller_id, org_id=org_id,
        authorization=authorization, db=db,
    )

    rows = result.get("datos") or result.get("por_taller") or []
    return rows, titulo


def _generate_csv(rows: list, col_keys: list, col_labels: list, titulo: str) -> io.StringIO:
    buf = io.StringIO()
    buf.write(f"# {titulo}\n")
    buf.write(f"# Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    if rows:
        writer = csv.writer(buf)
        writer.writerow(col_labels)
        for row in rows:
            writer.writerow([row.get(k, "") for k in col_keys])
    return buf


def _generate_excel(rows: list, col_keys: list, col_labels: list, titulo: str) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    n_cols = max(len(col_keys), 1)
    last_col = get_column_letter(n_cols)

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="5CBDB9", end_color="5CBDB9", fill_type="solid")
    title_font  = Font(bold=True, size=13, color="0F172A")

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"].value = titulo
    ws["A1"].font  = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].value = f"Asistencia Vehicular — Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font  = Font(italic=True, color="64748B", size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.row_dimensions[3].height = 6

    for col_idx, label in enumerate(col_labels, 1):
        cell = ws.cell(row=4, column=col_idx, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 18

    for row_idx, row in enumerate(rows, 5):
        for col_idx, key in enumerate(col_keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key))

    # auto-width based on data rows only (skip merged header rows)
    for col_cells in ws.iter_cols(min_row=4, max_col=n_cols):
        col_letter = col_cells[0].column_letter
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 35)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _generate_pdf(rows: list, col_keys: list, col_labels: list, titulo: str) -> io.BytesIO:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf = io.BytesIO()
    page = landscape(A4)
    margin = 1.5 * cm
    doc = SimpleDocTemplate(buf, pagesize=page,
                             topMargin=margin, bottomMargin=margin,
                             leftMargin=margin, rightMargin=margin)

    styles = getSampleStyleSheet()
    teal  = colors.HexColor("#5CBDB9")
    dark  = colors.HexColor("#0F172A")
    gray  = colors.HexColor("#64748B")
    light = colors.HexColor("#F8FAFC")

    story = []

    story.append(Paragraph(
        "Asistencia Vehicular",
        ParagraphStyle("brand", parent=styles["Normal"], textColor=gray,
                       fontSize=9, spaceAfter=2, alignment=TA_CENTER)
    ))
    story.append(Paragraph(
        titulo,
        ParagraphStyle("rep_title", parent=styles["Heading1"], textColor=dark,
                       fontSize=16, spaceAfter=2, alignment=TA_CENTER)
    ))
    story.append(Paragraph(
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle("sub", parent=styles["Normal"], textColor=gray,
                       fontSize=8, spaceAfter=10, alignment=TA_CENTER)
    ))

    if not rows:
        story.append(Paragraph("No hay datos para mostrar.", styles["Normal"]))
        doc.build(story)
        buf.seek(0)
        return buf

    # Column widths: wider for text columns, narrower for numbers/dates
    usable_w = page[0] - 2 * margin
    WIDE_COLS  = {'cliente', 'taller', 'vehiculo', 'tecnico', 'comentario'}
    NARROW_COLS = {'prioridad', 'estado', 'estado_asignacion', 'metodo_pago',
                   'completados', 'pendientes', 'urgentes', 'total', 'total_servicios'}
    col_widths = []
    for k in col_keys:
        if k in WIDE_COLS:
            col_widths.append(3.5)
        elif k in NARROW_COLS:
            col_widths.append(2.0)
        elif 'fecha' in k:
            col_widths.append(2.2)
        elif 'monto' in k or 'ingreso' in k or 'pct' in k or 'prom' in k or 'ticket' in k:
            col_widths.append(2.4)
        else:
            col_widths.append(2.8)

    total_weight = sum(col_widths)
    col_widths_pt = [usable_w * w / total_weight for w in col_widths]

    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, leading=10)

    table_data = [col_labels]
    for row in rows:
        table_data.append([
            Paragraph(str(row.get(k) or "—"), cell_style) for k in col_keys
        ])

    t = Table(table_data, colWidths=col_widths_pt, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0),  teal),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0),  8),
        ("ALIGN",        (0, 0), (-1, 0),  "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, light]),
        ("GRID",         (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
        ("LINEBELOW",    (0, 0), (-1, 0),  1, teal),
    ]))
    story.append(t)

    doc.build(story)
    buf.seek(0)
    return buf


@router.get("/exportar")
def exportar_reporte(
    tipo: str = Query(..., description="emergencias|historial-servicios|ingresos|calificaciones|kpis|incidentes-tipo|sla"),
    formato: str = Query(..., description="csv|pdf|excel"),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    taller_id: Optional[int] = Query(None),
    org_id: Optional[int] = Query(None),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    rows, titulo = _build_export_data(tipo, fecha_desde, fecha_hasta, taller_id, org_id, authorization, db)
    return _stream_export(rows, titulo, formato, tipo)


@router.post("/exportar-dinamico")
def exportar_dinamico(
    filtros: ReporteDinamicoFiltros,
    formato: str = Query(..., description="csv|pdf|excel"),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    result = reporte_dinamico(filtros, authorization=authorization, db=db)
    rows = result.get("datos", [])
    return _stream_export(rows, "Reporte Dinámico", formato, "dinamico")


def _stream_export(rows: list, titulo: str, formato: str, tipo: str) -> StreamingResponse:
    filtered_rows, col_keys, col_labels = _filter_rows(rows, tipo)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename_base = f"reporte_{tipo}_{ts}"

    if formato == "csv":
        buf = _generate_csv(filtered_rows, col_keys, col_labels, titulo)
        content = buf.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
        )

    if formato == "excel":
        buf = _generate_excel(filtered_rows, col_keys, col_labels, titulo)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
        )

    if formato == "pdf":
        buf = _generate_pdf(filtered_rows, col_keys, col_labels, titulo)
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.pdf"'},
        )

    raise HTTPException(status_code=400, detail="Formato no soportado. Use: csv, pdf, excel")
