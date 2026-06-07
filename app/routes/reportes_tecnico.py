from fastapi import APIRouter, HTTPException, Header, Query, Depends
from fastapi.responses import StreamingResponse
from psycopg2.extras import RealDictCursor
from pydantic import BaseModel
from typing import Optional
from datetime import datetime
import json
import csv
import io

from ..classes.postgresql import Database
from ..utils.tenant_deps import get_token_payload

router = APIRouter(prefix="/api/reportes/tecnico", tags=["Reportes Técnico"])


# ===================== HELPERS =====================

def _assert_tecnico(payload: dict) -> int:
    """Verifica que el token sea de un técnico y retorna su tecnico_id."""
    if payload.get("rol") != "tecnico":
        raise HTTPException(status_code=403, detail="Solo técnicos pueden acceder a estos reportes")
    tecnico_id = payload.get("tecnico_id")
    if not tecnico_id:
        raise HTTPException(status_code=403, detail="Token inválido: falta tecnico_id")
    return int(tecnico_id)


def _add_date_filter(conditions, params, fecha_desde=None, fecha_hasta=None, col="i.fecha_creacion"):
    if fecha_desde:
        conditions.append(f"{col} >= %s")
        params.append(fecha_desde)
    if fecha_hasta:
        conditions.append(f"{col} <= %s")
        params.append(fecha_hasta + " 23:59:59")


def _rows_to_dicts(rows) -> list:
    result = []
    for row in rows:
        d = {}
        for k, v in row.items():
            if hasattr(v, "isoformat"):
                d[k] = v.isoformat()
            elif v is None:
                d[k] = None
            else:
                d[k] = v
        result.append(d)
    return result


# ===================== MODELOS =====================

class VozTecnicoRequest(BaseModel):
    texto: str


# ===================== ENDPOINTS =====================

@router.get("/emergencias")
def tecnico_emergencias(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    """Emergencias atendidas por el técnico autenticado."""
    payload = get_token_payload(authorization)
    tecnico_id = _assert_tecnico(payload)

    conditions = ["a.tecnico_id = %s"]
    params = [tecnico_id]
    _add_date_filter(conditions, params, fecha_desde, fecha_hasta)

    where = " WHERE " + " AND ".join(conditions)
    query = """
        SELECT
            i.incidente_id,
            i.tipo_problema,
            i.descripcion,
            i.estado AS estado_incidente,
            i.prioridad,
            i.fecha_creacion,
            i.fecha_cierre,
            a.asignacion_id,
            a.estado AS estado_asignacion,
            a.fecha_asignacion,
            a.fecha_inicio_servicio,
            a.fecha_cierre_servicio,
            COALESCE(EXTRACT(EPOCH FROM (a.fecha_cierre_servicio - a.fecha_asignacion))/60, 0)::int AS duracion_min,
            t.razon_social AS taller,
            u.nombre AS cliente,
            u.telefono AS cliente_telefono,
            COALESCE(v.marca || ' ' || v.modelo, 'Sin vehículo') AS vehiculo,
            v.placa,
            COALESCE(p.monto_total, 0) AS monto_total
        FROM incidente i
        JOIN asignacion a ON a.incidente_id = i.incidente_id
        JOIN taller t ON t.taller_id = a.taller_id
        JOIN usuario u ON u.usuario_id = i.usuario_id
        LEFT JOIN vehiculo v ON v.vehiculo_id = i.vehiculo_id
        LEFT JOIN pago p ON p.incidente_id = i.incidente_id
    """ + where + " ORDER BY i.fecha_creacion DESC"

    cur = db.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(query, params)
        rows = cur.fetchall()
    finally:
        cur.close()

    data = _rows_to_dicts(rows)
    completadas = sum(1 for r in data if r.get("estado_asignacion") == "completada")
    monto_total = sum(r.get("monto_total") or 0 for r in data)

    return {
        "resumen": {
            "total": len(data),
            "completadas": completadas,
            "monto_total": round(monto_total, 2),
        },
        "datos": data,
    }


@router.get("/historial-servicios")
def tecnico_historial(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    """Historial de servicios completados por el técnico autenticado."""
    payload = get_token_payload(authorization)
    tecnico_id = _assert_tecnico(payload)

    conditions = ["a.tecnico_id = %s"]
    params = [tecnico_id]
    _add_date_filter(conditions, params, fecha_desde, fecha_hasta)

    where = " WHERE " + " AND ".join(conditions)
    query = """
        SELECT
            a.asignacion_id,
            i.incidente_id,
            i.tipo_problema,
            i.descripcion,
            a.estado,
            i.prioridad,
            i.fecha_creacion,
            a.fecha_asignacion,
            a.fecha_inicio_servicio,
            a.fecha_cierre_servicio,
            COALESCE(EXTRACT(EPOCH FROM (a.fecha_cierre_servicio - a.fecha_inicio_servicio))/60, 0)::int AS duracion_servicio_min,
            t.razon_social AS taller,
            u.nombre AS cliente,
            COALESCE(v.marca || ' ' || v.modelo, 'Sin vehículo') AS vehiculo,
            v.placa,
            COALESCE(p.monto_total, 0) AS monto_total,
            COALESCE(cal.puntuacion, 0) AS calificacion
        FROM asignacion a
        JOIN incidente i ON i.incidente_id = a.incidente_id
        JOIN taller t ON t.taller_id = a.taller_id
        JOIN usuario u ON u.usuario_id = i.usuario_id
        LEFT JOIN vehiculo v ON v.vehiculo_id = i.vehiculo_id
        LEFT JOIN pago p ON p.incidente_id = i.incidente_id
        LEFT JOIN calificacion cal ON cal.incidente_id = i.incidente_id
    """ + where + " ORDER BY a.fecha_asignacion DESC"

    cur = db.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(query, params)
        rows = cur.fetchall()
    finally:
        cur.close()

    data = _rows_to_dicts(rows)
    monto = sum(r.get("monto_total") or 0 for r in data)
    cal_list = [r["calificacion"] for r in data if r.get("calificacion")]
    promedio_cal = round(sum(cal_list) / len(cal_list), 2) if cal_list else 0

    return {
        "resumen": {
            "total_servicios": len(data),
            "ingresos_totales": round(monto, 2),
            "calificacion_promedio": promedio_cal,
        },
        "datos": data,
    }


@router.get("/calificaciones")
def tecnico_calificaciones(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    """Calificaciones recibidas por el técnico autenticado."""
    payload = get_token_payload(authorization)
    tecnico_id = _assert_tecnico(payload)

    conditions = ["a.tecnico_id = %s"]
    params = [tecnico_id]
    _add_date_filter(conditions, params, fecha_desde, fecha_hasta, col="c.fecha_calificacion")

    where = " WHERE " + " AND ".join(conditions)
    query = """
        SELECT
            c.calificacion_id,
            c.puntuacion,
            c.puntuacion_servicio,
            c.aspecto_atencion,
            c.aspecto_puntualidad,
            c.aspecto_limpieza,
            c.comentario,
            c.fecha_calificacion,
            t.razon_social AS taller,
            u.nombre AS cliente,
            i.tipo_problema
        FROM calificacion c
        JOIN incidente i ON i.incidente_id = c.incidente_id
        JOIN asignacion a ON a.incidente_id = i.incidente_id
        JOIN taller t ON t.taller_id = c.taller_id
        JOIN usuario u ON u.usuario_id = c.usuario_id
    """ + where + " ORDER BY c.fecha_calificacion DESC"

    cur = db.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(query, params)
        rows = cur.fetchall()
    finally:
        cur.close()

    data = _rows_to_dicts(rows)
    puntuaciones = [r["puntuacion"] for r in data if r.get("puntuacion")]
    promedio = round(sum(puntuaciones) / len(puntuaciones), 2) if puntuaciones else 0

    return {
        "resumen": {
            "total_calificaciones": len(data),
            "promedio_general": promedio,
            "distribucion": {str(i): sum(1 for p in puntuaciones if p == i) for i in range(1, 6)},
        },
        "datos": data,
    }


@router.get("/kpis")
def tecnico_kpis(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    """KPIs personales del técnico autenticado."""
    payload = get_token_payload(authorization)
    tecnico_id = _assert_tecnico(payload)

    conditions = ["a.tecnico_id = %s"]
    params = [tecnico_id]
    _add_date_filter(conditions, params, fecha_desde, fecha_hasta)

    where = " WHERE " + " AND ".join(conditions)
    query = f"""
        SELECT
            COUNT(a.asignacion_id) AS total_servicios,
            SUM(CASE WHEN a.estado = 'completada' THEN 1 ELSE 0 END) AS completados,
            SUM(CASE WHEN a.estado = 'pendiente' THEN 1 ELSE 0 END) AS pendientes,
            SUM(CASE WHEN i.prioridad = 'urgente' THEN 1 ELSE 0 END) AS urgentes,
            COALESCE(AVG(EXTRACT(EPOCH FROM (a.fecha_inicio_servicio - a.fecha_asignacion))/60), 0)::int AS tiempo_respuesta_prom,
            COALESCE(AVG(cal.puntuacion), 0)::numeric(3,1) AS calificacion_promedio
        FROM asignacion a
        JOIN incidente i ON i.incidente_id = a.incidente_id
        LEFT JOIN calificacion cal ON cal.incidente_id = i.incidente_id
        {where}
    """

    cur = db.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(query, params)
        row = cur.fetchone()
    finally:
        cur.close()

    data = _rows_to_dicts([row])[0] if row else {}
    total = int(data.get("total_servicios") or 0)
    completados = int(data.get("completados") or 0)
    tasa = round(completados / total * 100, 1) if total else 0

    return {
        "resumen": {
            "total_servicios": total,
            "completados": completados,
            "tasa_completacion_pct": tasa,
            "tiempo_respuesta_prom": int(data.get("tiempo_respuesta_prom") or 0),
            "calificacion_promedio": float(data.get("calificacion_promedio") or 0),
        },
        "datos": [data],
    }


@router.get("/incidentes-tipo")
def tecnico_incidentes_tipo(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    """Incidentes agrupados por tipo para el técnico autenticado."""
    payload = get_token_payload(authorization)
    tecnico_id = _assert_tecnico(payload)

    conditions = ["a.tecnico_id = %s"]
    params = [tecnico_id]
    _add_date_filter(conditions, params, fecha_desde, fecha_hasta)

    where = " WHERE " + " AND ".join(conditions)
    query = f"""
        SELECT
            COALESCE(i.tipo_problema, 'Sin clasificar') AS tipo_problema,
            COUNT(*) AS total,
            SUM(CASE WHEN a.estado = 'completada' THEN 1 ELSE 0 END) AS completados,
            SUM(CASE WHEN i.prioridad = 'urgente' THEN 1 ELSE 0 END) AS urgentes,
            ROUND(AVG(p.monto_total)::numeric, 2) AS ticket_promedio
        FROM incidente i
        JOIN asignacion a ON a.incidente_id = i.incidente_id
        LEFT JOIN pago p ON p.incidente_id = i.incidente_id
        {where}
        GROUP BY i.tipo_problema
        ORDER BY total DESC
    """

    cur = db.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(query, params)
        rows = cur.fetchall()
    finally:
        cur.close()

    return {"datos": _rows_to_dicts(rows)}


# ===================== REPORTE POR VOZ =====================

_INTENT_PROMPT_TECNICO = """Eres un asistente de reportes para un TÉCNICO de una plataforma de asistencia vehicular.
El técnico ha dicho (en texto): "{texto}"

Identifica el tipo de reporte y extrae filtros. Responde ÚNICAMENTE en JSON con este esquema:
{{
  "tipo_reporte": "emergencias|historial-servicios|calificaciones|kpis|incidentes-tipo",
  "filtros": {{
    "fecha_desde": "YYYY-MM-DD o null",
    "fecha_hasta": "YYYY-MM-DD o null"
  }},
  "mensaje_confirmacion": "frase corta confirmando lo que vas a generar"
}}

Reglas para inferir fechas relativas (hoy es {hoy}):
- "último mes" → fecha_desde = primer día del mes pasado, fecha_hasta = hoy
- "esta semana" → fecha_desde = lunes de esta semana
- "hoy" → fecha_desde = hoy
- "último trimestre" → fecha_desde = hace 90 días

Si no puedes determinar el tipo de reporte, usa "historial-servicios".
No incluyas ningún texto fuera del JSON."""


@router.post("/voz")
def tecnico_voz(
    body: VozTecnicoRequest,
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    """Reporte por voz/texto para el técnico autenticado."""
    from ..services.config import Config
    import google.generativeai as genai

    payload = get_token_payload(authorization)
    tecnico_id = _assert_tecnico(payload)

    try:
        genai.configure(api_key=Config.GEMINI_API_KEY)
        model = genai.GenerativeModel("gemini-2.5-flash")
        hoy = datetime.now().strftime("%Y-%m-%d")
        prompt = _INTENT_PROMPT_TECNICO.format(texto=body.texto, hoy=hoy)
        response = model.generate_content(prompt)
        raw = response.text.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        interpretacion = json.loads(raw)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo interpretar la consulta: {str(e)}")

    tipo = interpretacion.get("tipo_reporte", "historial-servicios")
    filtros_raw = interpretacion.get("filtros", {})
    mensaje = interpretacion.get("mensaje_confirmacion", "Generando reporte...")
    fecha_desde = filtros_raw.get("fecha_desde")
    fecha_hasta = filtros_raw.get("fecha_hasta")

    dispatch = {
        "emergencias": lambda: tecnico_emergencias(
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            authorization=authorization, db=db
        ),
        "historial-servicios": lambda: tecnico_historial(
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            authorization=authorization, db=db
        ),
        "calificaciones": lambda: tecnico_calificaciones(
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            authorization=authorization, db=db
        ),
        "kpis": lambda: tecnico_kpis(
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            authorization=authorization, db=db
        ),
        "incidentes-tipo": lambda: tecnico_incidentes_tipo(
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            authorization=authorization, db=db
        ),
    }

    resultado = dispatch.get(tipo, dispatch["historial-servicios"])()

    return {
        "tipo_reporte": tipo,
        "mensaje_confirmacion": mensaje,
        "filtros_aplicados": {
            "fecha_desde": fecha_desde,
            "fecha_hasta": fecha_hasta,
        },
        "resultado": resultado,
    }


# ===================== EXPORTACIÓN =====================

COLUMNAS_TECNICO = {
    "emergencias":         ["fecha_creacion", "tipo_problema", "estado_asignacion", "prioridad", "cliente", "vehiculo", "taller", "duracion_min", "monto_total"],
    "historial-servicios": ["fecha_asignacion", "tipo_problema", "estado", "prioridad", "cliente", "vehiculo", "taller", "duracion_servicio_min", "monto_total", "calificacion"],
    "calificaciones":      ["fecha_calificacion", "taller", "cliente", "tipo_problema", "puntuacion", "puntuacion_servicio", "comentario"],
    "kpis":                ["total_servicios", "completados", "pendientes", "urgentes", "tiempo_respuesta_prom", "calificacion_promedio"],
    "incidentes-tipo":     ["tipo_problema", "total", "completados", "urgentes", "ticket_promedio"],
}

COL_LABELS = {
    "fecha_creacion": "Fecha", "fecha_asignacion": "Fecha", "fecha_calificacion": "Fecha",
    "tipo_problema": "Tipo Incidente", "estado_asignacion": "Estado", "estado": "Estado",
    "prioridad": "Prioridad", "cliente": "Cliente", "vehiculo": "Vehículo", "taller": "Taller",
    "duracion_min": "Duración (min)", "duracion_servicio_min": "Duración (min)",
    "monto_total": "Monto Total", "calificacion": "Calificación",
    "puntuacion": "Puntuación", "puntuacion_servicio": "Punt. Servicio", "comentario": "Comentario",
    "total_servicios": "Total Servicios", "completados": "Completados", "pendientes": "Pendientes",
    "urgentes": "Urgentes", "tiempo_respuesta_prom": "T. Respuesta (min)",
    "calificacion_promedio": "Cal. Promedio", "total": "Total", "ticket_promedio": "Ticket Prom.",
}


def _prep_export(tipo: str, rows: list):
    cols = COLUMNAS_TECNICO.get(tipo, list(rows[0].keys()) if rows else [])
    if rows:
        available = set(rows[0].keys())
        cols = [c for c in cols if c in available]
        filtered = [{c: r.get(c) for c in cols} for r in rows]
    else:
        filtered = []
    labels = [COL_LABELS.get(c, c.replace("_", " ").title()) for c in cols]
    return filtered, cols, labels


@router.get("/exportar")
def tecnico_exportar(
    tipo: str = Query(..., description="emergencias|historial-servicios|calificaciones|kpis|incidentes-tipo"),
    formato: str = Query(..., description="csv|pdf|excel"),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    authorization: str = Header(None),
    db=Depends(Database.get_db),
):
    """Exporta el reporte del técnico en CSV, Excel o PDF."""
    payload = get_token_payload(authorization)
    _assert_tecnico(payload)

    dispatch = {
        "emergencias": lambda: tecnico_emergencias(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, authorization=authorization, db=db),
        "historial-servicios": lambda: tecnico_historial(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, authorization=authorization, db=db),
        "calificaciones": lambda: tecnico_calificaciones(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, authorization=authorization, db=db),
        "kpis": lambda: tecnico_kpis(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, authorization=authorization, db=db),
        "incidentes-tipo": lambda: tecnico_incidentes_tipo(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, authorization=authorization, db=db),
    }

    if tipo not in dispatch:
        raise HTTPException(status_code=400, detail=f"Tipo desconocido: {tipo}")

    result = dispatch[tipo]()
    rows = result.get("datos") or result.get("por_taller") or []
    titulos = {
        "emergencias": "Emergencias Atendidas",
        "historial-servicios": "Historial de Servicios",
        "calificaciones": "Calificaciones Recibidas",
        "kpis": "KPIs Personales",
        "incidentes-tipo": "Incidentes por Tipo",
    }
    titulo = titulos.get(tipo, "Reporte Técnico")
    filtered, cols, labels = _prep_export(tipo, rows)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reporte_tecnico_{tipo}_{ts}"

    if formato == "csv":
        buf = io.StringIO()
        buf.write(f"# {titulo}\n# Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        writer = csv.writer(buf)
        writer.writerow(labels)
        for row in filtered:
            writer.writerow([row.get(c, "") for c in cols])
        content = buf.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )

    if formato == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = titulo[:31]
        n_cols = max(len(cols), 1)
        last_col = get_column_letter(n_cols)
        ws.merge_cells(f"A1:{last_col}1")
        ws["A1"].value = titulo
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"A2:{last_col}2")
        ws["A2"].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font = Font(italic=True, size=9, color="64748B")
        ws["A2"].alignment = Alignment(horizontal="center")
        for ci, label in enumerate(labels, 1):
            cell = ws.cell(row=4, column=ci, value=label)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5CBDB9", end_color="5CBDB9", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(filtered, 5):
            for ci, key in enumerate(cols, 1):
                ws.cell(row=ri, column=ci, value=row.get(key))
        for col_cells in ws.iter_cols(min_row=4, max_col=n_cols):
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 35)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )

    if formato == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.enums import TA_CENTER
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
        styles = getSampleStyleSheet()
        teal = colors.HexColor("#5CBDB9")
        story = [
            Paragraph(titulo, ParagraphStyle("t", parent=styles["Heading1"], fontSize=16, alignment=TA_CENTER, spaceAfter=4)),
            Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ParagraphStyle("s", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#64748B"), alignment=TA_CENTER, spaceAfter=10)),
        ]
        if filtered:
            cell_style = ParagraphStyle("c", parent=styles["Normal"], fontSize=8, leading=10)
            table_data = [labels] + [[Paragraph(str(row.get(c) or "—"), cell_style) for c in cols] for row in filtered]
            usable_w = landscape(A4)[0] - 3*cm
            col_w = usable_w / len(cols)
            t = Table(table_data, colWidths=[col_w]*len(cols), repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), teal),
                ("TEXTCOLOR", (0,0), (-1,0), colors.white),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,-1), 8),
                ("ALIGN", (0,0), (-1,-1), "CENTER"),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#E2E8F0")),
            ]))
            story.append(t)
        else:
            story.append(Paragraph("No hay datos para mostrar.", styles["Normal"]))
        doc.build(story)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'})

    raise HTTPException(status_code=400, detail="Formato no soportado. Use: csv, pdf, excel")